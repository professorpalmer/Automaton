from __future__ import annotations

import io

from openpyxl import Workbook

from harness.recon_engine import flatten_row, run_workbook, run_workbooks


def _xlsx() -> bytes:
    book = Workbook()
    fa = book.active
    fa.title = "FA"
    fa.append(["id", "name"])
    fa.append(["1", "Ann"])
    fa.append(["2", "Ben"])
    fa.append(["3", "Cara"])
    lyft = book.create_sheet("Lyft")
    lyft.append(["id", "status"])
    lyft.append(["1", "Ride"])
    lyft.append(["2", "Cancel"])
    lyft.append(["9", "Ride"])
    buf = io.BytesIO()
    book.save(buf)
    return buf.getvalue()


def test_cancels_do_not_match_and_access_stays_honest() -> None:
    result = run_workbook(_xlsx(), "month.xlsx")
    assert result["matched"] == 1
    assert result["non_matchable"] == 1
    assert result["exceptions"] >= 1
    notes = " ".join(result["notes"])
    assert "FormAssembly" in notes
    assert "SharePoint" in notes
    assert result["ledger"][0]["matched_to"]["status"] == "Ride"


def test_org_name_cancel_is_not_skipped() -> None:
    book = Workbook()
    fa = book.active
    fa.title = "Requests"
    fa.append(["id", "org"])
    fa.append(["1", "Cancel Corp"])
    rides = book.create_sheet("Rides")
    rides.append(["id", "city"])
    rides.append(["1", "Austin"])
    buf = io.BytesIO()
    book.save(buf)
    result = run_workbook(buf.getvalue(), "month.xlsx")
    assert result["matched"] == 1
    assert result["non_matchable"] == 0


def test_lyft_sheet_first_still_pairs() -> None:
    book = Workbook()
    lyft = book.active
    lyft.title = "Lyft"
    lyft.append(["id", "status"])
    lyft.append(["1", "Ride"])
    lyft.append(["2", "Cancel"])
    fa = book.create_sheet("FormAssembly")
    fa.append(["id", "name"])
    fa.append(["1", "Ann"])
    fa.append(["2", "Ben"])
    buf = io.BytesIO()
    book.save(buf)
    result = run_workbook(buf.getvalue(), "month.xlsx")
    assert result["matched"] == 1
    assert result["non_matchable"] == 1


def test_xls_is_refused() -> None:
    result = run_workbook(b"not-a-workbook", "old.xls")
    assert "xlsx" in " ".join(result["notes"]).lower()
    assert result["matched"] == 0


def test_two_csvs_match() -> None:
    result = run_workbooks(
        [
            (b"id,name\n1,Ann\n", "fa.csv"),
            (b"id,status\n1,Ride\n", "lyft.csv"),
        ]
    )
    assert result["matched"] == 1
    assert flatten_row(result["ledger"][0])["matched_status"] == "Ride"

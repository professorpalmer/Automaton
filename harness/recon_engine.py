from __future__ import annotations

"""Workbook recon. Sheets in, counts and a match out. No live systems."""

import csv
import io
import re
from typing import Any, Dict, List, Optional

CANCEL = re.compile(r"\b(cancel|cancelled|canceled|refund)\b", re.I)
IDISH = re.compile(r"^(id|request|ticket|ride|booking|email|va|name)\b", re.I)
STATUS_COL = re.compile(r"^(status|state|result|outcome)$", re.I)
FA_SHEET = re.compile(r"formassembly|\bfa\b", re.I)
RIDE_SHEET = re.compile(r"lyft|flix", re.I)


def run_workbook(payload: bytes, filename: str = "upload.xlsx") -> Dict[str, Any]:
    return run_workbooks([(payload, filename)])


def run_workbooks(blobs: List[tuple[bytes, str]]) -> Dict[str, Any]:
    notes = [
        "This run used the file you uploaded.",
        "It did not connect to FormAssembly, Lyft, FlixBus, or SharePoint.",
    ]
    tables: List[Dict[str, Any]] = []
    for payload, filename in blobs:
        name = (filename or "").lower()
        if name.endswith(".xls") and not name.endswith(".xlsx"):
            notes.append("This .xls file needs to be saved as .xlsx or .csv.")
            return _result([], [], [], [], notes)
        tables.extend(_tables(payload, filename))
    if not tables:
        return _result(tables, [], [], [], notes + ["The file had no readable rows."])
    if len(tables) < 2:
        notes.append("Need two sheets (or two tables) to match. Showing counts only.")
        return _result(tables, [], [], tables[0]["rows"] if tables else [], notes)
    left, right = _pair_sheets(tables)
    key = _shared_key(left["columns"], right["columns"])
    if not key:
        notes.append("No shared id column. Nothing was matched.")
        return _result(tables, [], left["rows"] + right["rows"], [], notes)
    notes.append(f"Matched on {key}. Cancels and refunds were not matched.")
    matched, exceptions, skipped = _match(left["rows"], right["rows"], key)
    return _result(tables, matched, exceptions, skipped, notes)


def flatten_row(row: dict) -> dict:
    out = {}
    partner = row.get("matched_to")
    for key, value in row.items():
        if key == "matched_to":
            continue
        out[key] = value
    if isinstance(partner, dict):
        for key, value in partner.items():
            out[f"matched_{key}"] = value
    return out


def _tables(payload: bytes, filename: str) -> List[Dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return [_csv_table(payload, filename)]
    return _xlsx_tables(payload)


def _csv_table(payload: bytes, filename: str) -> Dict[str, Any]:
    text = payload.decode("utf-8-sig", errors="replace")
    rows = list(csv.DictReader(io.StringIO(text)))
    columns = list(rows[0].keys()) if rows else []
    return {"name": filename or "csv", "columns": columns, "rows": rows}


def _xlsx_tables(payload: bytes) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    tables = []
    for sheet in book.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        columns = [str(cell).strip() if cell is not None else f"col{i}" for i, cell in enumerate(header)]
        rows = []
        for raw in rows_iter:
            if all(cell is None or str(cell).strip() == "" for cell in raw):
                continue
            rows.append({columns[i]: "" if raw[i] is None else str(raw[i]).strip() for i in range(len(columns))})
        tables.append({"name": sheet.title, "columns": columns, "rows": rows})
    return tables


def _pair_sheets(tables: List[Dict[str, Any]]) -> tuple[Dict[str, Any], Dict[str, Any]]:
    fa = next((table for table in tables if FA_SHEET.search(table.get("name") or "")), None)
    ride = next((table for table in tables if RIDE_SHEET.search(table.get("name") or "")), None)
    if fa is not None and ride is not None:
        return fa, ride
    return tables[0], tables[1]


def _shared_key(left: List[str], right: List[str]) -> Optional[str]:
    right_l = {col.lower(): col for col in right}
    ranked = []
    for col in left:
        other = right_l.get(col.lower())
        if not other:
            continue
        ranked.append((0 if IDISH.search(col) else 1, col))
    ranked.sort()
    return ranked[0][1] if ranked else None


def _match(left: List[dict], right: List[dict], key: str) -> tuple[List[dict], List[dict], List[dict]]:
    right_key = _key_name(right, key)
    left_key = _key_name(left, key)
    pool = []
    skipped = []
    for row in right:
        if _is_cancel(row):
            skipped.append(row)
        else:
            pool.append(row)
    matched = []
    exceptions = []
    taken = set()
    for row in left:
        if _is_cancel(row):
            skipped.append(row)
            continue
        value = (row.get(left_key) or "").strip().lower()
        found = None
        for index, other in enumerate(pool):
            if index in taken:
                continue
            if value and (other.get(right_key) or "").strip().lower() == value:
                found = other
                taken.add(index)
                break
        if found is not None:
            matched.append({**row, "matched_to": found})
        else:
            exceptions.append({**row, "reason": "unmatched"})
    for index, other in enumerate(pool):
        if index not in taken:
            exceptions.append({**other, "reason": "unmatched"})
    return matched, exceptions, skipped


def _key_name(rows: List[dict], key: str) -> str:
    if not rows:
        return key
    for name in rows[0]:
        if name.lower() == key.lower():
            return name
    return key


def _is_cancel(row: dict) -> bool:
    status_vals = [value for key, value in row.items() if STATUS_COL.search(key or "")]
    if status_vals:
        return any(CANCEL.search(str(value)) for value in status_vals)
    return False


def _result(
    tables: List[dict],
    matched: List[dict],
    exceptions: List[dict],
    skipped: List[dict],
    notes: List[str],
) -> Dict[str, Any]:
    return {
        "sheets": [{"name": table["name"], "rows": len(table["rows"]), "columns": table["columns"]} for table in tables],
        "matched": len(matched),
        "exceptions": len(exceptions),
        "non_matchable": len(skipped),
        "ledger": matched,
        "exception_rows": exceptions,
        "notes": notes,
    }
